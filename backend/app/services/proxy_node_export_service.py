"""
代理节点导出服务层

支持将节点列表导出为 CSV 和 Excel (.xlsx) 格式。
"""
import csv
import io
import logging
from typing import List

import openpyxl
from openpyxl.styles import Font

from app.models.proxy_node import ProxyNode

logger = logging.getLogger(__name__)

# 导出列名（与导入模板保持一致）
EXPORT_COLUMNS = [
    "ip", "port", "username", "password", "protocol",
    "relay_ip", "relay_port", "relay_protocol",
    "purchase_date", "purchase_price", "purchase_channel",
    "expire_date", "sale_customer", "sale_price",
    "status", "usage", "remark",
]


def _node_to_row(node: ProxyNode) -> dict:
    """将 ProxyNode ORM 对象转换为可导出的字典（所有值均为字符串）。"""
    return {
        "ip": node.ip or "",
        "port": str(node.port) if node.port is not None else "",
        "username": node.username or "",
        "password": node.password or "",
        "protocol": node.protocol or "",
        "relay_ip": node.relay_ip or "",
        "relay_port": str(node.relay_port) if node.relay_port is not None else "",
        "relay_protocol": node.relay_protocol or "",
        "purchase_date": node.purchase_date.strftime("%Y-%m-%d") if node.purchase_date else "",
        "purchase_price": str(node.purchase_price) if node.purchase_price is not None else "",
        "purchase_channel": node.purchase_channel or "",
        "expire_date": node.expire_date.strftime("%Y-%m-%d") if node.expire_date else "",
        "sale_customer": node.sale_customer or "",
        "sale_price": str(node.sale_price) if node.sale_price is not None else "",
        "status": node.status or "",
        "usage": node.usage or "",
        "remark": node.remark or "",
    }


def export_to_csv(nodes: List[ProxyNode]) -> bytes:
    """
    将节点列表序列化为 CSV 格式。

    - 使用 UTF-8 with BOM 编码（兼容 Excel 直接打开）
    - 包含所有业务字段列（与导入模板列名一致）
    - nodes 为空列表时，仍返回包含列名行的文件
    """
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=EXPORT_COLUMNS)
    writer.writeheader()

    for node in nodes:
        writer.writerow(_node_to_row(node))

    logger.info(f"export_to_csv: 导出 {len(nodes)} 条节点")
    # UTF-8 with BOM
    return output.getvalue().encode("utf-8-sig")


def export_to_excel(nodes: List[ProxyNode]) -> bytes:
    """
    将节点列表序列化为 Excel .xlsx 格式。

    - 使用 openpyxl 创建工作簿
    - 第一行为列名（加粗），从第二行开始写入数据
    - 包含所有业务字段列（与导入模板列名一致）
    - nodes 为空列表时，仍返回包含列名行的文件
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "proxy_nodes"

    # 写入列名行（加粗）
    bold_font = Font(bold=True)
    for col_idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold_font

    # 写入数据行
    for row_idx, node in enumerate(nodes, start=2):
        row_data = _node_to_row(node)
        for col_idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])

    logger.info(f"export_to_excel: 导出 {len(nodes)} 条节点")

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
